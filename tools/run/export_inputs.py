from __future__ import annotations

from pathlib import Path
from typing import Sequence

import pandas as pd
import sys

from openpyxl import load_workbook

from engine.contracts import load_contract_registry, tables_for_surface
from engine.debug import init_debug_state, debug_enabled, debug_log


# =============================================================================
# Paths
# =============================================================================

def resolve_workbook_path(argv: Sequence[str]) -> Path:
    if len(argv) > 1:
        return Path(argv[1]).resolve()

    repo_root = Path(__file__).resolve().parents[2]
    return (repo_root / "runtime" / "Planner.xlsm").resolve()


def resolve_project_root(script_path: Path) -> Path:
    return script_path.resolve().parents[2]


def resolve_runtime_input_dir(project_root: Path) -> Path:
    return project_root / "data" / "runtime" / "input"


def resolve_reference_dir(project_root: Path) -> Path:
    return project_root / "data" / "reference"


# =============================================================================
# Excel Table Extraction
# =============================================================================

def read_excel_table(path: Path, table_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb["Inputs"]

    if table_name not in ws.tables:
        raise ValueError(f"Table not found: {table_name}")

    table = ws.tables[table_name]
    ref = table.ref  # e.g. A1:D20

    data = ws[ref]
    rows = list(data)

    headers = [cell.value for cell in rows[0]]
    records = [[cell.value for cell in row] for row in rows[1:]]

    df = pd.DataFrame(records, columns=headers)
    df = df.dropna(how="all")

    return df


# =============================================================================
# CSV
# =============================================================================

def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_csv(df: pd.DataFrame, path: Path) -> Path:
    df.to_csv(path, index=False)
    return path


# =============================================================================
# Input shaping helpers
# =============================================================================

def _find_enabled_column(df: pd.DataFrame) -> str | None:
    """
    Detect the enabled column in a case-insensitive, whitespace-tolerant way.
    Returns the exact matching column name from the DataFrame, or None.
    """
    for col in df.columns:
        if str(col).strip().lower() == "enabled":
            return col
    return None


def _filter_enabled_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    If an enabled column exists, keep only rows with truthy enabled values.
    Otherwise return the DataFrame unchanged.

    Accepted enabled values:
        TRUE, true, 1, yes, y
    """
    enabled_col = _find_enabled_column(df)

    if enabled_col is None:
        return df

    mask = (
        df[enabled_col]
        .astype(str)
        .str.strip()
        .str.lower()
        .isin(["true", "1", "yes", "y"])
    )

    return df.loc[mask].copy()


def _project_to_contract_fields(df: pd.DataFrame, contract) -> pd.DataFrame:
    """
    Strictly project DataFrame to contract-defined fields only, in contract order.
    Missing columns are added as empty values so output schema is stable.
    """
    contract_fields = list(contract.fields.keys())

    for col in contract_fields:
        if col not in df.columns:
            df[col] = ""

    return df[contract_fields].copy()


# =============================================================================
# Contract-driven export
# =============================================================================

def export_input_tables(
    workbook_path: Path,
    input_dir: Path,
    state,
):
    registry = load_contract_registry(workbook_path.parents[1] / "contracts")
    input_contracts = tables_for_surface(registry, "input")

    written = []

    # ---------------------------------------------------------
    # Precompute valid company keys from company table FIRST
    # ---------------------------------------------------------
    valid_company_keys = None

    if "company" in input_contracts:
        company_contract = input_contracts["company"]
        company_excel_table = "tblCompany"

        company_raw = read_excel_table(workbook_path, company_excel_table)
        company_enabled = _filter_enabled_rows(company_raw)
        company_df = _project_to_contract_fields(company_enabled, company_contract)

        if "company_key" in company_df.columns:
            valid_company_keys = set(company_df["company_key"])

        if debug_enabled(state, 3):
            debug_log(
                state,
                f"[ingest:company] valid_company_keys={sorted(valid_company_keys) if valid_company_keys is not None else None}",
                level=3,
            )

    # ---------------------------------------------------------
    # Export all input tables
    # ---------------------------------------------------------
    for table_name, contract in input_contracts.items():
        excel_table = f"tbl{''.join([part.capitalize() for part in table_name.split('_')])}"

        if debug_enabled(state, 2):
            debug_log(
                state,
                f"[ingest] reading table\n"
                f"  contract={table_name}\n"
                f"  excel_table={excel_table}",
                level=2,
            )

        # ---------------------------------------------------------
        # Read full Excel table
        # ---------------------------------------------------------
        df_raw = read_excel_table(workbook_path, excel_table)

        if debug_enabled(state, 3):
            debug_log(
                state,
                f"[ingest:{table_name}] raw_columns={list(df_raw.columns)}",
                level=3,
            )

        # ---------------------------------------------------------
        # Filter disabled rows before projection
        # ---------------------------------------------------------
        before_rows = len(df_raw)
        df_enabled = _filter_enabled_rows(df_raw)
        after_filter_rows = len(df_enabled)

        if debug_enabled(state, 3):
            enabled_col = _find_enabled_column(df_raw)
            debug_log(
                state,
                f"[ingest:{table_name}] enabled_filter\n"
                f"  enabled_column={enabled_col}\n"
                f"  before={before_rows}\n"
                f"  after={after_filter_rows}",
                level=3,
            )

        # ---------------------------------------------------------
        # Project strictly to contract columns
        # ---------------------------------------------------------
        df = _project_to_contract_fields(df_enabled, contract)

        # ---------------------------------------------------------
        # Enforce referential integrity against active companies
        # (skip company itself; it is the source of truth)
        # ---------------------------------------------------------
        if (
            table_name != "company"
            and "company_key" in df.columns
            and valid_company_keys is not None
        ):
            before_fk_rows = len(df)
            df = df[df["company_key"].isin(valid_company_keys)].copy()
            after_fk_rows = len(df)

            if debug_enabled(state, 3):
                debug_log(
                    state,
                    f"[ingest:{table_name}] company_fk_filter\n"
                    f"  before={before_fk_rows}\n"
                    f"  after={after_fk_rows}",
                    level=3,
                )

        if debug_enabled(state, 3):
            debug_log(
                state,
                f"[ingest:{table_name}] projected_columns={list(df.columns)}",
                level=3,
            )

        if debug_enabled(state, 2):
            debug_log(
                state,
                f"[ingest] rows\n"
                f"  table={table_name}\n"
                f"  rows={len(df)}",
                level=2,
            )

        if debug_enabled(state, 3) and not df.empty:
            debug_log(
                state,
                f"[ingest:{table_name}] sample\n"
                + "\n".join(str(r) for r in df.head(5).to_dict(orient="records")),
                level=3,
            )

        out_path = input_dir / f"{table_name}.csv"
        write_csv(df, out_path)

        written.append(out_path)

    return written


# =============================================================================
# Main
# =============================================================================

def main(argv: Sequence[str]) -> int:
    script_path = Path(__file__).resolve()
    project_root = resolve_project_root(script_path)

    workbook_path = resolve_workbook_path(argv)
    input_dir = ensure_directory(resolve_runtime_input_dir(project_root))
    ref_dir = resolve_reference_dir(project_root)

    state = init_debug_state(ref_dir)

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        return 1

    if debug_enabled(state, 1):
        debug_log(
            state,
            f"[ingest] reading workbook: {workbook_path}",
            level=1,
        )

    try:
        written_paths = export_input_tables(workbook_path, input_dir, state)
    except Exception as exc:
        print(str(exc))
        return 1

    if not written_paths:
        if debug_enabled(state, 1):
            debug_log(state, "[ingest] no input tables found", level=1)
        return 0

    if debug_enabled(state, 1):
        for path in written_paths:
            debug_log(state, f"[ingest] wrote: {path}", level=1)

        debug_log(state, "[ingest] export complete", level=1)

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))